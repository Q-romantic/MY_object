# 合并多个sheet表
def sheetmerge(filePath, newFilePath='', sheetname='汇总'):
    import pandas as pd
    sheetnames = pd.ExcelFile(filePath).sheet_names
    data = []
    for name in sheetnames:
        sheet_data = pd.read_excel(filePath, sheet_name=name)
        sheet_data.insert(0, '原表名', [name for i in range(sheet_data.shape[0])])  # sheet表索引放开头
        data.append(sheet_data)
        print(data)
    all_data = pd.concat(data, ignore_index=True)
    if newFilePath == '':
        from openpyxl import load_workbook
        book = load_workbook(filePath)
        writer = pd.ExcelWriter(filePath, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        all_data.to_excel(writer, sheet_name=sheetname, index=False)
        writer.save()
        writer.close()
    else:
        if '.xlsx' in str(newFilePath):
            all_data.to_excel(newFilePath, sheet_name=sheetname, index=False)
        else:
            newFilePath = filePath.replace('.', '_{}.').format(newFilePath)
            all_data.to_excel(newFilePath, sheet_name=sheetname, index=False)


# filePath = '销售清单.xlsx'
# newFilePath = '123销售清单.xlsx'
# # sheetmerge(filePath, newFilePath)  # 指定合并后文件名
# sheetmerge(filePath, 123)  # 指定合并后文件名加标识后缀
# # sheetmerge(filePath)  # 默认原表增加汇总表，不推荐


def workbook_merge_1(pwd, name):  # 注意：xlrd==1.2.1
    import pandas as pd
    import os
    new_path = f"{pwd}\\{name}.xlsx"
    if os.path.exists(new_path):
        os.remove(new_path)
    df_list = []
    for path, dirs, files in os.walk(pwd):
        for file in files:
            file_path = os.path.join(path, file)
            df = pd.read_excel(file_path)
            df_list.append(df)
    result = pd.concat(df_list)
    # print(result)
    result.to_excel(new_path, index=False)


pwd = "C:\\Y\\Case\\接任务练习\\code"
name = "tmp"
workbook_merge_1(pwd, name)


# 合并多个工作簿
def workbook_merge_2(pwd, name):
    import xlwt
    import xlrd
    import os
    new_path = f"{pwd}\\{name}.xlsx"
    if os.path.exists(new_path):
        os.remove(new_path)
    work = xlwt.Workbook(new_path)  # 建立一个文件
    file_list = os.listdir(pwd)
    for file in file_list:  # 循环遍历列出所有文件名称
        file_name = pwd + "\\" + file  # 路径+文件名
        workbook = xlrd.open_workbook(file_name)  # 打开第一个文件，注意：xlrd==1.2.0
        sheet_name = workbook.sheet_names()  # 获取第一个文件的sheet名称
        for file_1 in sheet_name:  # 循环遍历每个sheet
            val = []
            sheet = work.add_sheet(file_1, cell_overwrite_ok=True)  # 新建一个sheet
            table = workbook.sheet_by_name(file_1)  # 以名字为索引
            rows = table.nrows  # 获取sheet行数
            clos = table.ncols  # 获取sheet列数目
            for i in range(rows):  # 循环遍历没一行
                val.append(table.row_values(i))  # 获取没一行的值
                for x in range(len(val)):
                    for y in range(len(val[x])):
                        sheet.write(x, y, val[x][y])
    work.save(new_path)


# pwd = 'C:\\Y\\Case\\接任务练习\\code'  # 目录(放Excel表格的目录)
# name = "tmp"  # 合成后工作簿文件名
# workbook_merge_2(pwd, name)

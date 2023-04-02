# -*- coding: utf-8 -*-
"""
@Time    : 2023/2/1  001 上午 10:12
@Author  : Jan
@File    : 20220201_下载超链接图片放表格.py
"""

""" {} """

import openpyxl
import glob
import os
import cv2
import requests
import numpy as np
import xlsxwriter
import threading
from openpyxl.utils import get_column_letter as gcl  # 3 -> 'C'
from openpyxl.utils import column_index_from_string  # ('A' -> 1)

try:
    for i in glob.glob("*.jpg"):
        os.remove(i)
except:
    pass

try:
    os.remove("tmp.xlsx")
except:
    pass

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 Edg/109.0.1518.70"}


def get_url_and_download_jpg(row, column):
    main_book = openpyxl.load_workbook("场部问题明细报表2023-01-30-201.xlsx")
    main_sheet = main_book.active
    print(main_sheet.cell(row=row, column=column).value)
    print(main_sheet.cell(row=row, column=column).hyperlink.target)  # 读取超链接
    url = main_sheet.cell(row=row, column=column).hyperlink.target
    # print(url)
    # print(url.split('?')[-1])
    url2 = f"https://ym-chain.wens.com.cn/v1/chain/patrol/pictureShow?{url.split('?')[-1]}"
    response = requests.get(url2, headers=headers)
    data = response.json()
    # print(data)
    url3 = data['data']['picUrl']
    # print(url3)
    response = requests.get(url3, headers=headers)
    data = response.content
    # # print(data)
    # print(response)
    path = f'{column}-{row}.jpg'
    # with open(path, mode='wb') as f:
    #     f.write(data)

    data = np.asarray(bytearray(data), dtype="uint8")
    img = cv2.imdecode(data, cv2.IMREAD_COLOR)
    y, x = img.shape[0:2]  # 高，宽
    cv2.imwrite(path, img)

    return path, x, y, url


def main(row, column, index):
    path, x, y, url = get_url_and_download_jpg(row, column)
    sheet.write(f"{gcl(index)}{row}", f"{gcl(column)}-{row}", my_format)
    sheet.write_url(f"{gcl(index + 1)}{row}", url, string='url-link-alias', tip='Click here', cell_format=None)
    sheet.insert_image(f"{gcl(index + 2)}{row + 0})", path, {"x_scale": round(68 / x, 4), "y_scale": round(16 / y, 4), "x_offset": 2, "y_offset": 2})


# 可选重复或带参数，多线程并发
def start_thread(functon, *args, **kwargs):
    """
    start_thread(func=func, num=2)  # 重复执行func()用法
    start_thread(func=func, params=i)   # 传参执行func(i)用法
    :param kwargs:
    :return:
    """

    def run_func(*args, **kwargs):
        return functon(*args) if args else functon()

    for i in range(kwargs['num'] if kwargs.get('num') else 1):
        thread_name = threading.Thread(target=run_func, args=args, kwargs=kwargs) if args else threading.Thread(target=run_func)
        thread_name.start()


if __name__ == '__main__':
    book = xlsxwriter.Workbook(r"tmp.xlsx")
    sheet = book.add_worksheet()
    my_format = book.add_format({
        'bold': True,  # 字体加粗
        'align': 'center',  # 水平位置设置：居中
        'valign': 'vcenter',  # 垂直位置设置，居中
        'font_size': 12,  # 字体大小设置
        'font_color': 'red',  # 字体颜色
        'underline': True,  # 下换线
    })

    columns = ['W', 'Y', 'AA', 'AC', 'AO']  # 需要下载图片的列号
    columns = [column_index_from_string(i) for i in columns]
    for column, index in zip(columns, range(1, 4 * len(columns), 4)):  # 每组间隔一个空列
        for row in range(2, 40):  # 涉及到的行数
            try:
                # start_thread(main, row, column, index)  没能保存结果？？？
                main(row, column, index)
            except:
                pass
        #     break
        # break
    book.close()

    try:
        for i in glob.glob("*.jpg"):
            os.remove(i)
    except:
        pass
